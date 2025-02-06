import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from PIL import Image as Im

fp =  r"C:\Users\mtsullivan\OneDrive - DOI\In_Depth_Analysis\ND_Test\inputs\graphs"

arr = {}

new_size = (400, 250)

"""
for filename in os.listdir(fp):
    if filename.lower().endswith(".png"):  # Only process PNG files
        file_path = os.path.join(fp, filename)
        
        try:
            # Open the image
            with Im.open(file_path) as img:
                # Resize the image
                img_resized = img.resize(new_size, Im.LANCZOS)
                
                # Save back to the same path (overwrite)
                img_resized.save(file_path, optimize = True)
        except Exception as e:
            print(f"Error processing {filename}: {e}")
"""

# Organize images into dictionary based on keys
for file in os.listdir(fp):
    x = os.path.join(fp, file)
    if ".PNG" in x:
        num_part = x.split('\\')[-1].split('_')[0]
        if num_part not in arr:
            arr[num_part] = []
        updated_path = x.replace(r"graphs", r"op_graphs")
        arr[num_part].append(updated_path)

        with Im.open(x) as img:
            # Downsize to the target size (150x75)
            img_resized = img.resize(new_size, Im.LANCZOS)

            # Save the resized image back to the same path (overwrite)
            img_resized.save(updated_path, format="PNG", optimize=True)


max_length = float('-inf')  # Start with the smallest possible value

# Loop through the dictionary
for key, val in arr.items():
    local_max = len(val)  # Find the max in the current list
    if local_max > max_length:
        max_length = local_max


fp_part = fp.split('\\')[-1]


# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "Images"

# Set header row
ws.append(['ID', 'Image'])

# Starting row for images
start_row = 2

# Loop through the dictionary
for ID, filepaths in arr.items():
    # Insert the ID into the row
    ws.cell(row=start_row, column=1, value=ID)
    amount = len(filepaths)
    count = 0
    # Inserting images into the next column
    for i, filepath in enumerate(filepaths):
        try:
            if count == 0:
                # Load the image and add it to the worksheet
                img = Image(filepath)
                # Calculate the position for the image
                # We will position it according to the row and column
                img_anchor = f'B{start_row + i}'  # Adjusting for multiple images in the same ID row
                ws.add_image(img, img_anchor)
            elif count == 1:
                # Load the image and add it to the worksheet
                img = Image(filepath)
                # Calculate the position for the image
                # We will position it according to the row and column
                img_anchor = f'I{start_row + i}'  # Adjusting for multiple images in the same ID row
                ws.add_image(img, img_anchor)
            elif count == 2:
                # Load the image and add it to the worksheet
                img = Image(filepath)
                # Calculate the position for the image
                # We will position it according to the row and column
                img_anchor = f'Q{start_row + i}'  # Adjusting for multiple images in the same ID row
                ws.add_image(img, img_anchor)
            elif count == 3:
                # Load the image and add it to the worksheet
                img = Image(filepath)
                # Calculate the position for the image
                # We will position it according to the row and column
                img_anchor = f'Y{start_row + i}'  # Adjusting for multiple images in the same ID row
                ws.add_image(img, img_anchor)
            elif count == 4:
                # Load the image and add it to the worksheet
                img = Image(filepath)
                # Calculate the position for the image
                # We will position it according to the row and column
                img_anchor = f'AG{start_row + i}'  # Adjusting for multiple images in the same ID row
                ws.add_image(img, img_anchor)
            elif count == 5:
                # Load the image and add it to the worksheet
                img = Image(filepath)
                # Calculate the position for the image
                # We will position it according to the row and column
                img_anchor = f'AO{start_row + i}'  # Adjusting for multiple images in the same ID row
                ws.add_image(img, img_anchor)
            count += 1
        except Exception as e:
            print(f"Could not add image for {ID}: {filepath} - {e}")

    start_row += 18  # Move to the next row for the next ID

# Save the workbook
output_fp = r'C:\Users\mtsullivan\OneDrive - DOI\In_Depth_Analysis\ND_Test\inputs\graphs\image_excel3.xlsx'
wb.save(output_fp)

# Print confirmation message
print(f"Excel file created with images at: {output_fp}")
